import base64
import html
import io
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import pandas as pd
import streamlit as st

st.set_page_config(page_title="ExcelClean AI", page_icon="📊", layout="wide")

ESTILO = """
<style>
@import url("https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&display=swap");

html, body, .stApp {
  font-family: "IBM Plex Sans", sans-serif;
}

.stApp {
  background:
    linear-gradient(105deg, rgba(11, 18, 32, 0.94) 0%, rgba(11, 18, 32, 0.82) 42%, rgba(11, 18, 32, 0.55) 100%),
    url("data:image/jpeg;base64,__FONDO__") right center / cover no-repeat;
}

[data-testid="stHeader"] {
  background: rgba(11, 18, 32, 0.72);
  backdrop-filter: blur(10px);
}

[data-testid="stSidebar"] {
  background: rgba(14, 22, 38, 0.94);
  border-right: 1px solid rgba(232, 238, 246, 0.08);
}

[data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3 {
  letter-spacing: 0.04em;
}

.hero {
  border: 1px solid rgba(43, 184, 168, 0.28);
  background:
    radial-gradient(ellipse 70% 80% at 0% 0%, rgba(43,184,168,0.16), transparent 55%),
    linear-gradient(145deg, rgba(14, 22, 32, 0.94), rgba(10, 16, 24, 0.90));
  border-radius: 12px;
  padding: 0.85rem 1.1rem 0.8rem;
  margin-bottom: 0.75rem;
}

.hero-kicker {
  color: #2BB8A8;
  font-size: 0.68rem;
  font-weight: 700;
  letter-spacing: 0.16em;
  text-transform: uppercase;
  margin: 0 0 0.25rem 0;
}

.hero h1 {
  margin: 0;
  font-size: 1.35rem;
  font-weight: 700;
  color: #F4F7FB;
}

.hero p {
  margin: 0.3rem 0 0 0;
  color: #B7C3D4;
  font-size: 0.85rem;
  line-height: 1.35;
}

.vx-filter-title {
  margin: 0 0 0.15rem 0;
  font-size: 0.7rem;
  font-weight: 700;
  letter-spacing: 0.1em;
  text-transform: uppercase;
  color: #8A9BB0;
}

.vx-filter-heading {
  margin: 0 0 0.75rem 0;
  font-size: 1.08rem;
  font-weight: 700;
  color: #EAF0F6;
}

.sb-card-title {
  font-size: 0.7rem;
  letter-spacing: 0.1em;
  text-transform: uppercase;
  color: #8A9BB0;
  font-weight: 700;
  margin: 0 0 0.3rem 0;
}

.sb-card-heading {
  font-size: 1.08rem;
  font-weight: 700;
  color: #EAF0F6;
  margin: 0 0 0.85rem 0;
  line-height: 1.3;
}

[data-testid="stVerticalBlockBorderWrapper"] {
  background: #0C1219 !important;
  border: 1px solid #1A2636 !important;
  border-radius: 12px !important;
  padding: 1rem 1.15rem 1.2rem !important;
  box-shadow: 0 16px 48px rgba(0, 0, 0, 0.4);
}

[data-testid="stDataFrame"] {
  border: 1px solid #1A2636 !important;
  border-radius: 8px !important;
  overflow: hidden;
  background: #090E15 !important;
}

[data-testid="stFileUploader"] {
  border: 1px dashed rgba(43, 184, 168, 0.45);
  border-radius: 10px;
  padding: 0.35rem 0.55rem;
  background: rgba(12, 18, 25, 0.72);
}

[data-testid="stFileUploader"] section {
  border: none !important;
  padding: 0 !important;
}

[data-testid="stFileUploaderDropzone"] {
  min-height: 2.5rem !important;
  padding: 0.3rem 0.45rem !important;
}

div[data-testid="stMetric"] {
  background: rgba(18, 26, 42, 0.8);
  border: 1px solid rgba(232, 238, 246, 0.08);
  border-radius: 12px;
  padding: 0.55rem 0.75rem;
}

.stButton > button {
  border-radius: 10px;
  font-weight: 600;
  letter-spacing: 0.02em;
}

.stDownloadButton > button {
  border-radius: 10px;
  font-weight: 600;
}
</style>
"""


FONDO_PATH = Path(__file__).resolve().parent / "assets" / "fondo.jpg"


def aplicar_estilo() -> None:
    css = ESTILO
    if FONDO_PATH.exists():
        foto = base64.b64encode(FONDO_PATH.read_bytes()).decode("ascii")
        css = css.replace("__FONDO__", foto)
    st.markdown(css, unsafe_allow_html=True)

PALABRAS_NUMERO = (
    "monto",
    "precio",
    "total",
    "cantidad",
    "valor",
    "costo",
    "importe",
    "pago",
    "saldo",
)
CARACTERES_TEXTO = r"[^a-zA-Z0-9\sñÑáéíóúÁÉÍÓÚüÜ]"
MAX_MB = 25


def es_columna_numero(nombre: str) -> bool:
    n = str(nombre).lower()
    return any(k in n for k in PALABRAS_NUMERO)


def es_columna_fecha(nombre: str) -> bool:
    n = str(nombre).lower()
    return "fecha" in n or "date" in n


def normalizar_nombre_columna(nombre) -> str:
    texto = str(nombre).strip()
    texto = re.sub(r"\s+", " ", texto)
    if texto.lower().startswith("unnamed") or texto in ("", "nan", "None"):
        return ""
    return texto


def a_numero(serie: pd.Series) -> pd.Series:
    texto = serie.astype(str).str.strip()
    texto = texto.replace({"": pd.NA, "nan": pd.NA, "None": pd.NA, "NaT": pd.NA})
    # 1.234,56 → 1234.56  |  1234.56 se deja
    con_coma = texto.str.contains(",", na=False)
    texto = texto.where(~con_coma, texto.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    texto = texto.str.replace(r"[^\d.\-]", "", regex=True)
    texto = texto.replace({"": pd.NA})
    return pd.to_numeric(texto, errors="coerce")


def es_columna_id(nombre: str) -> bool:
    n = str(nombre).lower()
    return any(k in n for k in ("id", "codigo", "código", "sscc", "ean", "sku"))


def ids_a_texto(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if not es_columna_id(col):
            continue
        def _celda(v):
            if pd.isna(v):
                return ""
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            s = str(v).strip()
            if "e+" in s.lower() or "e-" in s.lower():
                try:
                    return str(int(float(s)))
                except ValueError:
                    return s
            if s.endswith(".0"):
                return s[:-2]
            return s
        out[col] = out[col].map(_celda)
    return out


def leer_hoja(datos: bytes, hoja: str) -> pd.DataFrame:
    tabla = pd.read_excel(io.BytesIO(datos), sheet_name=hoja, engine="openpyxl")
    return ids_a_texto(tabla)


def config_columnas(df: pd.DataFrame) -> dict:
    cfg = {}
    for col in df.columns:
        if es_columna_id(col):
            cfg[col] = st.column_config.TextColumn(str(col), width="medium")
    return cfg


def filtrar_revision(df: pd.DataFrame, busqueda: str, columna: str, valor: str) -> pd.DataFrame:
    vista = df
    if busqueda.strip():
        mask = vista.astype(str).apply(
            lambda row: row.str.contains(busqueda, case=False, na=False).any(),
            axis=1,
        )
        vista = vista[mask]
    if columna != "(todas)" and valor != "TODOS":
        vista = vista[vista[columna].astype(str) == str(valor)]
    return vista


def cuadro_filtro_validador(df: pd.DataFrame, clave: str) -> pd.DataFrame:
    st.markdown(
        '<p class="vx-filter-title">Revisión de lote</p>'
        '<p class="vx-filter-heading">Buscar y filtrar (Validador)</p>',
        unsafe_allow_html=True,
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        busqueda = st.text_input(
            "Búsqueda global (lote / producto / almacén)",
            key=f"{clave}_busq",
        )
    with c2:
        columna = st.selectbox(
            "Filtrar por columna",
            ["(todas)"] + list(map(str, df.columns)),
            key=f"{clave}_col",
        )
    with c3:
        if columna != "(todas)" and columna in df.columns:
            unicos = (
                df[columna]
                .dropna()
                .astype(str)
                .replace({"": pd.NA})
                .dropna()
                .unique()
                .tolist()
            )
            unicos = sorted(unicos)[:400]
            valor = st.selectbox("Valor", ["TODOS"] + unicos, key=f"{clave}_val")
        else:
            valor = "TODOS"
            st.selectbox("Valor", ["TODOS"], key=f"{clave}_val_off", disabled=True)
    vista = filtrar_revision(df, busqueda, columna, valor)
    st.caption(f"Mostrando **{len(vista):,}** de **{len(df):,}** filas")
    return vista


def editar_en_cuadro(df: pd.DataFrame, hoja: str, prefijo: str, guardar) -> pd.DataFrame:
    hoja_txt = html.escape(str(hoja))
    with st.container(border=True):
        st.markdown(
            f'<p class="sb-card-title">Excel · almacén / packing</p>'
            f'<p class="sb-card-heading">{hoja_txt}</p>',
            unsafe_allow_html=True,
        )
        df = herramientas_columnas(df, prefijo, guardar)
        vista = cuadro_filtro_validador(df, f"filtro_{prefijo}")
        filas = "fixed" if len(vista) != len(df) else "dynamic"
        editado_vista = editor_excel(vista, f"editor_{prefijo}", filas=filas)
        if len(vista) == len(df):
            guardar(editado_vista)
            return editado_vista
        base = df.copy()
        comunes = editado_vista.index.intersection(base.index)
        for col in editado_vista.columns:
            if col in base.columns:
                base.loc[comunes, col] = editado_vista.loc[comunes, col]
        guardar(base)
        return base


def limpiar_dataframe(
    df: pd.DataFrame,
    *,
    eliminar_duplicados: bool,
    eliminar_filas_vacias: bool,
    eliminar_columnas_vacias: bool,
    rellenar_na: bool,
    limpiar_espacios: bool,
    espacios_internos: bool,
    modo_texto: str,
    remover_especiales: bool,
    corregir_numeros: bool,
    corregir_fechas: bool,
    normalizar_encabezados: bool,
) -> tuple[pd.DataFrame, dict]:
    df_limpio = df.copy()
    stats = {
        "filas_antes": int(df.shape[0]),
        "cols_antes": int(df.shape[1]),
        "duplicados": 0,
        "filas_vacias": 0,
        "cols_vacias": 0,
    }

    if normalizar_encabezados:
        nuevos = [normalizar_nombre_columna(c) for c in df_limpio.columns]
        usados = {}
        finales = []
        for i, nombre in enumerate(nuevos):
            base = nombre or f"columna_{i + 1}"
            if base in usados:
                usados[base] += 1
                base = f"{base}_{usados[base]}"
            else:
                usados[base] = 1
            finales.append(base)
        df_limpio.columns = finales

    if eliminar_filas_vacias:
        antes = len(df_limpio)
        df_limpio = df_limpio.dropna(how="all")
        stats["filas_vacias"] = antes - len(df_limpio)

    if eliminar_columnas_vacias:
        antes = df_limpio.shape[1]
        df_limpio = df_limpio.dropna(axis=1, how="all")
        stats["cols_vacias"] = antes - df_limpio.shape[1]

    if eliminar_duplicados:
        antes = len(df_limpio)
        df_limpio = df_limpio.drop_duplicates()
        stats["duplicados"] = antes - len(df_limpio)

    for col in df_limpio.columns:
        serie = df_limpio[col]
        es_texto = serie.dtype == "object" or pd.api.types.is_string_dtype(serie)
        fecha = es_columna_fecha(col)
        numero = es_columna_numero(col)

        if es_texto and limpiar_espacios:
            mask = serie.notna()
            df_limpio.loc[mask, col] = serie.loc[mask].astype(str).str.strip()
            serie = df_limpio[col]

        if es_texto and espacios_internos:
            mask = df_limpio[col].notna()
            df_limpio.loc[mask, col] = (
                df_limpio.loc[mask, col].astype(str).str.replace(r"\s+", " ", regex=True)
            )

        if es_texto and modo_texto != "dejar" and not fecha:
            mask = df_limpio[col].notna()
            s = df_limpio.loc[mask, col].astype(str)
            if modo_texto == "MAYÚSCULAS":
                df_limpio.loc[mask, col] = s.str.upper()
            elif modo_texto == "minúsculas":
                df_limpio.loc[mask, col] = s.str.lower()
            elif modo_texto == "Título":
                df_limpio.loc[mask, col] = s.str.title()

        if es_texto and remover_especiales and not fecha and not numero:
            mask = df_limpio[col].notna()
            df_limpio.loc[mask, col] = (
                df_limpio.loc[mask, col].astype(str).str.replace(CARACTERES_TEXTO, "", regex=True)
            )

        if corregir_numeros and numero:
            df_limpio[col] = a_numero(df_limpio[col])

        if corregir_fechas and fecha:
            df_limpio[col] = pd.to_datetime(df_limpio[col], errors="coerce", dayfirst=True).dt.strftime(
                "%Y-%m-%d"
            )

    if rellenar_na:
        df_limpio = df_limpio.fillna("N/A")

    stats["filas_despues"] = int(df_limpio.shape[0])
    stats["cols_despues"] = int(df_limpio.shape[1])
    return df_limpio, stats


AZUL_OSCURO = "1F4E78"
AZUL_CLARO = "D6E3F0"
ROJO_VACIO = "FFCCCC"
BLANCO = "FFFFFF"
TEXTO_EXCEL = "1A1A1A"


def _borde_excel():
    thin = Side(style="thin", color="B0B0B0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def aplicar_estilo_validador(ws, titulo: str, n_cols: int, n_filas: int) -> None:
    if n_cols < 1:
        return
    azul = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
    zebra = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type="solid")
    rojo = PatternFill(start_color=ROJO_VACIO, end_color=ROJO_VACIO, fill_type="solid")
    font_titulo = Font(name="Calibri", bold=True, color=BLANCO, size=14)
    font_head = Font(name="Calibri", bold=True, color=BLANCO, size=11)
    font_dato = Font(name="Calibri", color=TEXTO_EXCEL, size=10)
    borde = _borde_excel()
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    celda_t = ws.cell(row=1, column=1, value=titulo)
    celda_t.font = font_titulo
    celda_t.fill = azul
    celda_t.alignment = centro
    ws.row_dimensions[1].height = 28
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = azul
        cell.border = borde

    ws.row_dimensions[2].height = 22
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = font_head
        cell.fill = azul
        cell.alignment = centro
        cell.border = borde

    ultima = 2 + max(n_filas, 0)
    for r in range(3, ultima + 1):
        fila_clara = r % 2 == 0
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_dato
            cell.alignment = izq
            cell.border = borde
            val = cell.value
            if val is None or str(val).strip() in ("", "-", "nan"):
                cell.fill = rojo
            elif fila_clara:
                cell.fill = zebra

    for c in range(1, n_cols + 1):
        letra = get_column_letter(c)
        max_len = 10
        for r in range(2, ultima + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            for parte in str(val).splitlines() or [""]:
                max_len = max(max_len, len(parte))
        ws.column_dimensions[letra].width = min(max(max_len + 3, 12), 60)

    ws.freeze_panes = "A3"
    ws.print_title_rows = "1:2"


def excel_en_memoria(hojas: dict[str, pd.DataFrame]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for nombre, tabla in hojas.items():
        hoja = re.sub(r"[\\/*?:\[\]]", "_", str(nombre))[:31] or "limpio"
        df_out = tabla.copy() if tabla is not None else pd.DataFrame()
        if df_out.empty and df_out.columns.empty:
            df_out = pd.DataFrame({"Aviso": ["Sin datos"]})
        ws = wb.create_sheet(title=hoja)
        n_cols = max(len(df_out.columns), 1)
        for c, col_name in enumerate(df_out.columns, start=1):
            ws.cell(row=2, column=c, value=str(col_name))
        for r_idx, row in enumerate(df_out.itertuples(index=False), start=3):
            for c_idx, valor in enumerate(row, start=1):
                if pd.isna(valor):
                    valor = ""
                ws.cell(row=r_idx, column=c_idx, value=valor)
        titulo = f"Packing List — ExcelClean AI | {nombre}"
        aplicar_estilo_validador(ws, titulo, n_cols, len(df_out))
    if not wb.sheetnames:
        ws = wb.create_sheet("limpio")
        aplicar_estilo_validador(ws, "Packing List — ExcelClean AI", 1, 0)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def borrar_editores() -> None:
    for k in list(st.session_state.keys()):
        if str(k).startswith("editor_"):
            del st.session_state[k]


def editor_excel(df: pd.DataFrame, clave: str, altura: int = 520, filas="dynamic") -> pd.DataFrame:
    return st.data_editor(
        df,
        num_rows=filas,
        use_container_width=True,
        height=altura,
        key=clave,
        hide_index=True,
        column_config=config_columnas(df),
    )


def herramientas_columnas(df: pd.DataFrame, prefijo: str, guardar) -> pd.DataFrame:
    c1, c2, c3 = st.columns([2, 1, 2])
    with c1:
        nueva = st.text_input(
            "Nueva columna",
            placeholder="Ej: OBSERVACION",
            key=f"{prefijo}_nueva_col",
        )
    with c2:
        st.write("")
        st.write("")
        if st.button("Agregar", key=f"{prefijo}_add_col") and nueva.strip():
            nombre = nueva.strip()
            if nombre not in df.columns:
                df = df.copy()
                df[nombre] = ""
                guardar(df)
                st.session_state.editor_ver = st.session_state.get("editor_ver", 0) + 1
                borrar_editores()
                st.rerun()
    with c3:
        if len(df.columns):
            borrar = st.selectbox(
                "Quitar columna",
                ["(ninguna)"] + list(map(str, df.columns)),
                key=f"{prefijo}_del_col",
            )
            if borrar != "(ninguna)" and st.button("Quitar", key=f"{prefijo}_del_btn"):
                df = df.drop(columns=[borrar])
                guardar(df)
                st.session_state.editor_ver = st.session_state.get("editor_ver", 0) + 1
                borrar_editores()
                st.rerun()
    return df


aplicar_estilo()

st.sidebar.markdown(
    """
    <div style="padding:0 0 0.7rem 0;margin-bottom:0.35rem;border-bottom:1px solid rgba(43,184,168,0.22)">
      <p style="margin:0;font-size:0.68rem;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;color:#2BB8A8">Validador</p>
      <p style="margin:0.12rem 0 0 0;font-size:0.82rem;font-weight:600;color:#EAF0F6">ExcelClean AI</p>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
<div class="hero">
  <p class="hero-kicker">Validador · packing y almacén</p>
  <h1>ExcelClean AI</h1>
</div>
""",
    unsafe_allow_html=True,
)

archivo = st.file_uploader("Archivo Excel (.xlsx)", type=["xlsx"])

if archivo is None:
    for k in ("resultado", "archivo_nombre", "hojas_elegidas", "edit_src", "editor_ver"):
        st.session_state.pop(k, None)
    borrar_editores()
    st.stop()

peso_mb = len(archivo.getvalue()) / (1024 * 1024)
if peso_mb > MAX_MB:
    st.error(f"El archivo pesa {peso_mb:.1f} MB. El máximo es {MAX_MB} MB.")
    st.stop()

if st.session_state.get("archivo_nombre") != archivo.name:
    st.session_state.archivo_nombre = archivo.name
    st.session_state.pop("resultado", None)
    st.session_state.pop("edit_src", None)
    st.session_state.editor_ver = 0
    borrar_editores()

datos = archivo.getvalue()

try:
    xls = pd.ExcelFile(io.BytesIO(datos), engine="openpyxl")
    hojas = xls.sheet_names
except Exception as e:
    st.error(f"Error al leer el Excel: {e}")
    st.stop()

st.sidebar.header("Hojas")
if len(hojas) == 1:
    hojas_elegidas = hojas
    st.sidebar.caption(f"Una hoja: **{hojas[0]}**")
else:
    todas = st.sidebar.checkbox("Limpiar todas las hojas", value=False)
    if todas:
        hojas_elegidas = hojas
    else:
        hojas_elegidas = st.sidebar.multiselect(
            "Hojas a tabular", hojas, default=[hojas[0]]
        )
    if not hojas_elegidas:
        st.warning("Elige al menos una hoja.")
        st.stop()

if st.session_state.get("hojas_elegidas") != hojas_elegidas:
    st.session_state.hojas_elegidas = hojas_elegidas
    st.session_state.pop("resultado", None)

try:
    tablas = {h: leer_hoja(datos, h) for h in hojas_elegidas}
except Exception as e:
    st.error(f"Error al leer la hoja: {e}")
    st.stop()

if "edit_src" not in st.session_state:
    st.session_state.edit_src = {}
for h, tabla in tablas.items():
    if h not in st.session_state.edit_src:
        st.session_state.edit_src[h] = tabla.copy()

st.sidebar.header("Filtros")

st.sidebar.subheader("1. Filas y columnas")
eliminar_duplicados = st.sidebar.checkbox("Eliminar filas 100% idénticas", value=True)
eliminar_filas_vacias = st.sidebar.checkbox(
    "Eliminar filas completamente vacías", value=True
)
eliminar_columnas_vacias = st.sidebar.checkbox(
    "Eliminar columnas completamente vacías", value=True
)
normalizar_encabezados = st.sidebar.checkbox(
    "Limpiar nombres de columnas", value=True
)
rellenar_na = st.sidebar.checkbox("Rellenar celdas vacías sueltas con N/A")

st.sidebar.subheader("2. Texto")
limpiar_espacios = st.sidebar.checkbox("Quitar espacios al inicio/final", value=True)
espacios_internos = st.sidebar.checkbox("Dejar un solo espacio entre palabras", value=True)
modo_texto = st.sidebar.selectbox(
    "Mayúsculas / minúsculas",
    ["dejar", "MAYÚSCULAS", "minúsculas", "Título"],
)
remover_especiales = st.sidebar.checkbox("Quitar caracteres raros (#, $, %, @)")

st.sidebar.subheader("3. Números y fechas")
corregir_numeros = st.sidebar.checkbox(
    "Forzar números (monto, precio, total, importe)", value=True
)
corregir_fechas = st.sidebar.checkbox(
    "Normalizar fechas a YYYY-MM-DD (día primero)", value=True
)

if st.sidebar.button("Limpiar ahora", use_container_width=True, type="primary"):
    with st.spinner("Tabulando y limpiando..."):
        resultado = {}
        for nombre in hojas_elegidas:
            tabla = st.session_state.edit_src.get(nombre, tablas[nombre])
            limpio, stats = limpiar_dataframe(
                tabla,
                eliminar_duplicados=eliminar_duplicados,
                eliminar_filas_vacias=eliminar_filas_vacias,
                eliminar_columnas_vacias=eliminar_columnas_vacias,
                rellenar_na=rellenar_na,
                limpiar_espacios=limpiar_espacios,
                espacios_internos=espacios_internos,
                modo_texto=modo_texto,
                remover_especiales=remover_especiales,
                corregir_numeros=corregir_numeros,
                corregir_fechas=corregir_fechas,
                normalizar_encabezados=normalizar_encabezados,
            )
            resultado[nombre] = {"original": tabla, "limpio": limpio, "stats": stats}
        st.session_state.resultado = resultado
        st.session_state.editor_ver = st.session_state.get("editor_ver", 0) + 1
        borrar_editores()

resultado = st.session_state.get("resultado")
ver = st.session_state.get("editor_ver", 0)

st.subheader("Tabla de revisión")
st.caption("El Excel queda dentro del mismo cuadro de lote que usa Validador.")
hoja_vista = (
    hojas_elegidas[0]
    if len(hojas_elegidas) == 1
    else st.selectbox("Ver hoja", hojas_elegidas)
)

if resultado is None:
    df = st.session_state.edit_src[hoja_vista]
    c1, c2, c3 = st.columns(3)
    c1.metric("Filas", f"{df.shape[0]:,}")
    c2.metric("Columnas", f"{df.shape[1]:,}")
    c3.metric("Hojas", len(hojas_elegidas))
    editado = editar_en_cuadro(
        df,
        hoja_vista,
        f"src_{hoja_vista}_{ver}",
        lambda t: st.session_state.edit_src.__setitem__(hoja_vista, t),
    )
    st.info("Revisa el lote en el cuadro o pulsa **Limpiar ahora**.")
    st.download_button(
        label="Descargar esta hoja",
        data=excel_en_memoria({hoja_vista: editado}),
        file_name=f"{Path(archivo.name).stem}_{hoja_vista}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.stop()

item = resultado[hoja_vista]
stats = item["stats"]

m1, m2, m3, m4 = st.columns(4)
m1.metric("Filas listas", f"{item['limpio'].shape[0]:,}", delta=-(stats["filas_antes"] - item["limpio"].shape[0]))
m2.metric("Duplicados quitados", stats["duplicados"])
m3.metric("Filas vacías quitadas", stats["filas_vacias"])
m4.metric("Columnas vacías quitadas", stats["cols_vacias"])

antes, despues = st.tabs(["Original (editable)", "Limpio (editable)"])
with antes:
    orig = editar_en_cuadro(
        item["original"],
        hoja_vista,
        f"orig_{hoja_vista}_{ver}",
        lambda t: st.session_state.resultado[hoja_vista].__setitem__("original", t),
    )
    st.session_state.edit_src[hoja_vista] = orig
with despues:
    editar_en_cuadro(
        item["limpio"],
        hoja_vista,
        f"limpio_{hoja_vista}_{ver}",
        lambda t: st.session_state.resultado[hoja_vista].__setitem__("limpio", t),
    )

nombre_salida = f"{Path(archivo.name).stem}_limpio.xlsx"
st.download_button(
    label="Descargar Excel limpio",
    data=excel_en_memoria({n: r["limpio"] for n, r in st.session_state.resultado.items()}),
    file_name=nombre_salida,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
