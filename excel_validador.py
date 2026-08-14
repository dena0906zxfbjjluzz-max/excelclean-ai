import io
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AZUL_OSCURO = "1F4E78"
AZUL_CLARO = "D6E3F0"
BLANCO = "FFFFFF"
TEXTO_EXCEL = "1A1A1A"


def _borde_excel():
    thin = Side(style="thin", color="B0B0B0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def aplicar_estilo_tabla(ws, n_cols: int, n_filas: int) -> None:
    if n_cols < 1:
        return
    azul = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
    zebra = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type="solid")
    font_head = Font(name="Calibri", bold=True, color=BLANCO, size=11)
    font_dato = Font(name="Calibri", color=TEXTO_EXCEL, size=10)
    borde = _borde_excel()
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 22
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = font_head
        cell.fill = azul
        cell.alignment = centro
        cell.border = borde

    ultima = 1 + max(n_filas, 0)
    for r in range(2, ultima + 1):
        fila_clara = r % 2 == 0
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_dato
            cell.alignment = izq
            cell.border = borde
            if fila_clara:
                cell.fill = zebra

    for c in range(1, n_cols + 1):
        letra = get_column_letter(c)
        max_len = 10
        for r in range(1, ultima + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            for parte in str(val).splitlines() or [""]:
                max_len = max(max_len, len(parte))
        ws.column_dimensions[letra].width = min(max(max_len + 3, 12), 60)

    ws.freeze_panes = "A2"
    ws.print_title_rows = "1:1"


def excel_en_memoria(hojas: dict[str, pd.DataFrame]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for nombre, tabla in hojas.items():
        hoja = re.sub(r"[\\/*?:\[\]]", "_", str(nombre))[:31] or "limpio"
        df_out = tabla.copy() if tabla is not None else pd.DataFrame()
        if df_out.empty and df_out.columns.empty:
            df_out = pd.DataFrame({"Aviso": ["Sin datos"]})
        ws = wb.create_sheet(title=hoja)
        n_cols = max(len(df_out.columns), 1)
        for c, col_name in enumerate(df_out.columns, start=1):
            ws.cell(row=1, column=c, value=str(col_name))
        for r_idx, row in enumerate(df_out.itertuples(index=False), start=2):
            for c_idx, valor in enumerate(row, start=1):
                if pd.isna(valor):
                    valor = ""
                ws.cell(row=r_idx, column=c_idx, value=valor)
        aplicar_estilo_tabla(ws, n_cols, len(df_out))
    if not wb.sheetnames:
        ws = wb.create_sheet("limpio")
        aplicar_estilo_tabla(ws, 1, 0)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
